"""Tenant-scoped, authenticated report generation and downloads."""

import os
from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
from typing import Literal, Optional
from uuid import uuid4
from xml.sax.saxutils import escape

import openpyxl
from beanie import PydanticObjectId
from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import FileResponse
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, Field, model_validator
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.models.models import Report, Transaction, User
from app.services.auth import require_report_access


router = APIRouter(prefix="/reports", tags=["reports"])

ReportType = Literal["transaction_summary", "aml_risk_report", "shariat_audit"]
ExportFormat = Literal["pdf", "excel"]
TransactionType = Literal["Murabaha", "Musharaka"]
TransactionStatus = Literal["pending", "reviewing", "approved", "rejected"]

PDF_CONTENT_TYPE = "application/pdf"
EXCEL_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
MAX_REPORT_BYTES = 12 * 1024 * 1024
MAX_REPORT_ROWS = 5000


class ReportGenerateRequest(BaseModel):
    report_type: ReportType
    export_format: ExportFormat
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    transaction_type: Optional[TransactionType] = None
    status: Optional[TransactionStatus] = None
    min_amount: Optional[float] = Field(default=None, ge=0)

    @model_validator(mode="after")
    def validate_date_range(self):
        if self.start_date and self.end_date and self.start_date > self.end_date:
            raise ValueError("Boshlanish sanasi tugash sanasidan keyin bo'lishi mumkin emas")
        return self


async def filter_transactions(
    req: ReportGenerateRequest,
    tenant_id: str,
) -> list[Transaction]:
    filters: dict = {"tenant_id": tenant_id}
    if req.transaction_type:
        filters["type"] = req.transaction_type
    if req.status:
        filters["status"] = req.status
    if req.min_amount is not None:
        filters["amount"] = {"$gte": req.min_amount}
    if req.start_date or req.end_date:
        created_at_filter: dict = {}
        if req.start_date:
            created_at_filter["$gte"] = datetime.combine(req.start_date, time.min)
        if req.end_date:
            created_at_filter["$lte"] = datetime.combine(req.end_date, time.max)
        filters["created_at"] = created_at_filter

    return await Transaction.find(filters).sort("-created_at").limit(MAX_REPORT_ROWS + 1).to_list()


def _totals_by_currency(txs: list[Transaction]) -> dict[str, float]:
    totals: dict[str, float] = {}
    for tx in txs:
        totals[tx.currency] = totals.get(tx.currency, 0) + tx.amount
    return totals


def _excel_safe(value: Optional[str]) -> str:
    """Prevent user-controlled cells from being interpreted as formulas."""
    text = value or "—"
    if text.startswith(("=", "+", "-", "@", "\t", "\r")):
        return f"'{text}"
    return text


def generate_pdf_report(
    req: ReportGenerateRequest,
    txs: list[Transaction],
    bank_name: str,
) -> bytes:
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "DocTitle",
        parent=styles["Heading1"],
        fontSize=20,
        textColor=colors.HexColor("#1B4332"),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "DocSubtitle",
        parent=styles["Normal"],
        fontSize=11,
        textColor=colors.HexColor("#C9A227"),
        spaceAfter=15,
    )
    report_titles = {
        "transaction_summary": "BITIMLAR XULOSASI HISOBOTI",
        "aml_risk_report": "AML / KYC RISK TAHLILI HISOBOTI",
        "shariat_audit": "SHARIAT KENGASHI AUDIT HISOBOTI",
    }

    story.append(
        Paragraph(
            f"<b>MIZAN</b> — {report_titles[req.report_type]}",
            title_style,
        )
    )
    story.append(
        Paragraph(
            f"Tashkilot: {escape(bank_name)} | "
            f"Yaratilgan vaqt: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            subtitle_style,
        )
    )
    story.append(
        HRFlowable(
            width="100%",
            thickness=1.5,
            color=colors.HexColor("#C9A227"),
            spaceAfter=15,
        )
    )

    totals = _totals_by_currency(txs)
    summary_data = [["Jami bitimlar soni", f"{len(txs)} ta"]]
    summary_data.extend(
        [f"Umumiy summa ({currency})", f"{amount:,.0f}"]
        for currency, amount in sorted(totals.items())
    )
    summary_table = Table(summary_data, colWidths=[220, 220])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F4F6F4")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#0F2D21")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 15)])

    table_data = [["ID", "Tur", "Miqdor", "Kontragent", "Holat", "Risk", "Sana"]]
    for tx in txs:
        table_data.append(
            [
                tx.transaction_id,
                tx.type,
                f"{tx.amount:,.0f} {tx.currency}",
                (tx.counterparty or "—")[:20],
                tx.status.upper(),
                (tx.risk_score or "low").upper(),
                tx.created_at.strftime("%Y-%m-%d"),
            ]
        )

    tx_table = Table(table_data, colWidths=[50, 65, 100, 120, 75, 55, 65], repeatRows=1)
    tx_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B4332")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E0E0E0")),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#F9FAF9")],
                ),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.extend([tx_table, Spacer(1, 20)])
    footer_style = ParagraphStyle(
        "DocFooter",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.HexColor("#6B7D67"),
        alignment=1,
    )
    story.append(
        Paragraph(
            "Ushbu hisobot MIZAN platformasi tomonidan avtomatik yaratilgan.",
            footer_style,
        )
    )
    doc.build(story)
    return output.getvalue()


def generate_excel_report(
    req: ReportGenerateRequest,
    txs: list[Transaction],
    bank_name: str,
) -> bytes:
    del req  # The selected transactions already represent the request filters.
    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "Xulosa statistika"

    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1B4332")

    summary["A1"] = "MIZAN PLATFORMA HISOBOTI"
    summary["A1"].font = title_font
    summary["A2"] = (
        f"Tashkilot: {_excel_safe(bank_name)} | "
        f"Yaratilgan: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    summary.append([])
    summary.append(["Ko'rsatkich", "Qiymat"])
    for column in range(1, 3):
        cell = summary.cell(row=4, column=column)
        cell.fill = header_fill
        cell.font = header_font

    summary.append(["Jami bitimlar soni", len(txs)])
    for currency, amount in sorted(_totals_by_currency(txs).items()):
        summary.append([f"Umumiy hajm ({currency})", amount])
    summary.append(["Tasdiqlangan bitimlar", sum(tx.status == "approved" for tx in txs)])
    summary.append(["Rad etilgan bitimlar", sum(tx.status == "rejected" for tx in txs)])

    details = workbook.create_sheet(title="To'liq bitimlar ro'yxati")
    headers = [
        "ID",
        "Tur",
        "Miqdor",
        "Valyuta",
        "Mas'ul shaxs",
        "Kontragent",
        "Holat",
        "Risk score",
        "Yaratilgan sana",
    ]
    details.append(headers)
    for column in range(1, len(headers) + 1):
        cell = details.cell(row=1, column=column)
        cell.fill = header_fill
        cell.font = header_font

    for tx in txs:
        details.append(
            [
                tx.transaction_id,
                tx.type,
                tx.amount,
                tx.currency,
                _excel_safe(tx.responsible_person),
                _excel_safe(tx.counterparty),
                tx.status,
                tx.risk_score or "low",
                tx.created_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )

    for worksheet in (summary, details):
        worksheet.freeze_panes = "A2"
        for column in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 3, 12), 60)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _serialize_report(report: Report) -> dict:
    filename = report.filename
    if not filename and report.file_path:
        filename = Path(report.file_path).name
    return {
        "id": str(report.id),
        "report_type": report.report_type,
        "export_format": report.format,
        "created_by": report.created_by,
        "created_at": report.created_at,
        "filename": filename or "report",
        "download_url": f"/api/reports/download/{report.id}",
    }


@router.post("/generate", status_code=201)
async def generate_report(
    req: ReportGenerateRequest,
    current_user: User = Depends(require_report_access),
):
    txs = await filter_transactions(req, current_user.tenant_id)
    if len(txs) > MAX_REPORT_ROWS:
        raise HTTPException(
            status_code=413,
            detail="Hisobotda 5000 tadan ko'p bitim bor. Filtrlarni toraytiring.",
        )
    filename = (
        f"report_{req.report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_"
        f"{uuid4().hex[:8]}.{'pdf' if req.export_format == 'pdf' else 'xlsx'}"
    )
    bank_name = current_user.bank_name or "O'zbekiston Islom Banki"
    if req.export_format == "pdf":
        file_data = generate_pdf_report(req, txs, bank_name)
        content_type = PDF_CONTENT_TYPE
    else:
        file_data = generate_excel_report(req, txs, bank_name)
        content_type = EXCEL_CONTENT_TYPE

    if len(file_data) > MAX_REPORT_BYTES:
        raise HTTPException(
            status_code=413,
            detail="Hisobot juda katta. Filtrlarni toraytirib qayta urinib ko'ring.",
        )

    report = Report(
        tenant_id=current_user.tenant_id,
        report_type=req.report_type,
        format=req.export_format,
        filters=req.model_dump(mode="json", exclude_none=True),
        filename=filename,
        content_type=content_type,
        file_data=file_data,
        created_by=current_user.full_name,
    )
    await report.insert()
    record = _serialize_report(report)
    return {"id": str(report.id), "filename": filename, "record": record, **record}


@router.get("/history")
async def get_reports_history(
    current_user: User = Depends(require_report_access),
):
    reports = await Report.find(
        Report.tenant_id == current_user.tenant_id
    ).sort("-created_at").limit(50).to_list()
    return [_serialize_report(report) for report in reports]


@router.get("/download/{report_id}")
async def download_report(
    report_id: str,
    current_user: User = Depends(require_report_access),
):
    if not PydanticObjectId.is_valid(report_id):
        raise HTTPException(status_code=404, detail="Hisobot topilmadi")

    report = await Report.get(PydanticObjectId(report_id))
    if not report or report.tenant_id != current_user.tenant_id:
        raise HTTPException(status_code=404, detail="Hisobot topilmadi")

    filename = report.filename or (
        Path(report.file_path).name if report.file_path else "report"
    )
    content_type = report.content_type or (
        PDF_CONTENT_TYPE if report.format == "pdf" else EXCEL_CONTENT_TYPE
    )
    if report.file_data is not None:
        return Response(
            content=report.file_data,
            media_type=content_type,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if report.file_path and os.path.isfile(report.file_path):
        return FileResponse(
            path=report.file_path,
            filename=filename,
            media_type=content_type,
        )
    raise HTTPException(status_code=404, detail="Hisobot fayli topilmadi")
